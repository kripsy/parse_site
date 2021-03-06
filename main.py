import urllib.parse
import requests
from bs4 import BeautifulSoup
from lxml import html
from openpyxl import Workbook, load_workbook
from datetime import datetime
import csv

answer = []

description = ["Номер", "Наименование оператора", "ИНН",
               "ФИО физического лица или наименование юридического лица, ответственных за обработку персональных данных",
               "номера их контактных телефонов, почтовые адреса и адреса электронной почты"]


class Company:
    def __init__(self, number, name, inn, FIO, info):
        self.number = number
        self.name = name
        self.inn = inn
        self.FIO = FIO
        self.info = info


urlglobal = "http://rkn.gov.ru/personal-data/register/"


def get_html_id(url, name):
    # name = "ВТБ Банк Москвы"
    a = name#.encode('windows-1251')
    values = {'inn': a}
    data = urllib.parse.urlencode(values).encode('utf-8')
    r = requests.get(url, data)
    try:
        parse_id(r.text)
    except AttributeError:
        print(name)

    return


# получили ссылку на страницу компании
def parse_id(html):
    soup = BeautifulSoup(html)
    table = soup.find('table', class_='TblList1')
    ids = []
    for row in table.find_all('tr')[1:]:
        cols = row.find_all('td')[0]
        ids.append(cols.nobr.text)
    nexturl = []
    for x in range(len(ids)):
        nexturl.append(urlglobal + "?id=" + ids[x])
    bigdata = []
    # можно красивей, но сейчас 4 утра
    for x in nexturl:
        bigdata.append(get_html_info(x))

    for x in bigdata:
        answer.append(parse_info(x))
    # вот здесь будем запиливать все в файл excel
    f1 = open('text.txt', 'w')
    for x in answer:
        f1.write(x.number + ' ')
        f1.write(x.name + ' ')
        f1.write(x.inn + ' ')
        f1.write(x.FIO + ' ')
        f1.write(x.info + '\n')
    f1.close()
    return


# получили информацию о компании с ее страницы
def get_html_info(url):
    r = requests.get(url)
    return (r.text.encode('cp1251'))


def parse_info(html):
    soup = BeautifulSoup(html)
    table = soup.find('table', class_='TblList')
    info = []
    for row in table.find_all('tr'):
        col = row.find_all('td')
        if (col[0].text == description[0]):
            info.append(col[1].text)
        if (col[0].text == description[1]):
            info.append(col[1].text)
        if (col[0].text == description[2]):
            info.append(col[1].text)
        if (col[0].text == description[3]):
            info.append(col[1].text)
        if (col[0].text == description[4]):
            info.append(col[1].text)
    return (Company(info[0], info[1], info[2], info[3], info[4]))


def main():

    name_company = []
    # with open('rating.csv', 'rt', encoding='cp1251') as csvfile:
    #     reader = csv.reader(csvfile, delimiter = ';', quotechar='|')
    #     i = 0
    #     for row in reader:
    #         i += 1
    #         if i > 4:
    #             name_company.insert(len(name_company), row[2])
    #         if len(name_company) > 80:
    #             print(name_company)
    #             return

    wb = load_workbook(filename='company_inn_1.xlsx', read_only=True)
    ws = wb.active
    for row in ws.rows:
        for cell in row:
            print(cell.value)
            name_company.insert(len(name_company), cell.value)
    start(name_company)
    return


def start(name_company):
    for i in name_company:
        get_html_id(urlglobal, i)
    return


def check_to_xlsx():
    wb = Workbook(write_only=False)
    ws = wb.active
    ws['A1'] = description[0]
    ws['B1'] = description[1]
    ws['C1'] = description[2]
    ws['D1'] = description[3]
    ws['E1'] = description[4]
    i = 2
    for rec in range(len(answer)):
        ws.cell(row=i, column=1).value = answer[rec].number
        ws.cell(row=i, column=2).value = answer[rec].name
        ws.cell(row=i, column=3).value = answer[rec].inn
        ws.cell(row=i, column=4).value = answer[rec].FIO
        ws.cell(row=i, column=5).value = answer[rec].info
        i += 1
    wb.save('new_big_file.xlsx')
    return


if __name__ == '__main__':
    main()
    check_to_xlsx()
